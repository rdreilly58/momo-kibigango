#!/usr/bin/env python3
"""
xlsx.py — Read and write Microsoft Excel (.xlsx) files

Usage:
  xlsx.py read FILE.xlsx                         # List sheets
  xlsx.py read FILE.xlsx SHEET_NAME              # Read sheet as table
  xlsx.py read FILE.xlsx SHEET_NAME --json       # JSON output
  xlsx.py create "Sheet1" -o FILE.xlsx           # Create new workbook
  xlsx.py append FILE.xlsx SHEET_NAME HEADER=... # Append data
"""

import sys
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def read_xlsx(filepath, sheet_name=None, as_json=False):
    """Read Excel workbook."""
    wb = load_workbook(filepath)
    
    if sheet_name is None:
        # List all sheets
        return "\n".join([f"  {name}" for name in wb.sheetnames])
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    
    ws = wb[sheet_name]
    
    if as_json:
        data = {
            "sheet": sheet_name,
            "rows": []
        }
        for row in ws.iter_rows(values_only=True):
            data["rows"].append(list(row))
        return json.dumps(data, indent=2)
    else:
        # Table output
        lines = [f"Sheet: {sheet_name}"]
        lines.append("=" * 80)
        
        # Header row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").ljust(20))
        lines.append(" | ".join(headers))
        lines.append("-" * 80)
        
        # Data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            cols = [str(cell or "").ljust(20) for cell in row]
            lines.append(" | ".join(cols))
        
        return "\n".join(lines)


def create_xlsx(sheet_name, filepath, headers=None):
    """Create a new Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if headers:
        ws.append(headers)
    
    wb.save(filepath)
    print(f"[xlsx] Created: {filepath}")


def append_to_xlsx(filepath, sheet_name, data):
    """Append a row to an Excel sheet."""
    wb = load_workbook(filepath)
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    
    # Parse key=value pairs
    values = []
    for item in data:
        if "=" in item:
            values.append(item.split("=", 1)[1])
        else:
            values.append(item)
    
    ws.append(values)
    wb.save(filepath)
    print(f"[xlsx] Updated: {filepath}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    cmd = sys.argv[1]
    
    if cmd == "read":
        filepath = sys.argv[2]
        sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
        as_json = "--json" in sys.argv
        
        result = read_xlsx(filepath, sheet_name, as_json=as_json)
        print(result)
    
    elif cmd == "create":
        sheet_name = sys.argv[2] if len(sys.argv) > 2 else "Sheet1"
        output = None
        headers = None
        
        for i, arg in enumerate(sys.argv):
            if arg == "-o" and i + 1 < len(sys.argv):
                output = sys.argv[i + 1]
            elif arg == "-h" and i + 1 < len(sys.argv):
                headers = sys.argv[i + 1].split(",")
        
        if not output:
            print("Error: -o OUTPUT required")
            sys.exit(1)
        
        create_xlsx(sheet_name, output, headers)
    
    elif cmd == "append":
        filepath = sys.argv[2]
        sheet_name = sys.argv[3]
        data = sys.argv[4:]
        
        if not data:
            print("Error: no data to append")
            sys.exit(1)
        
        append_to_xlsx(filepath, sheet_name, data)
    
    else:
        print(f"Unknown command: {cmd}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
